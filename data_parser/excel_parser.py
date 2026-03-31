import os
import hashlib
import subprocess
import tempfile
from typing import List
from .base_parser import BaseParser, DataBlock, ModalityType


class ExcelParser(BaseParser):
    """Excel文件解析器

    将Excel文件解析为数据块，每个工作表合并为一个数据块。
    """

    def __init__(self, cache_dir: str = "cache/data_blocks"):
        super().__init__()
        self.supported_extensions = ['xlsx', 'xls']
        self.cache_dir = cache_dir

    def _get_cache_path(self, file_path: str) -> str:
        """生成缓存目录路径"""
        file_hash = hashlib.md5(os.path.abspath(file_path).encode()).hexdigest()[:12]
        file_name = os.path.splitext(os.path.basename(file_path))[0]

        cache_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            self.cache_dir,
            f"{file_name}_{file_hash}"
        )

        os.makedirs(cache_path, exist_ok=True)
        return cache_path

    def parse(self, file_path: str) -> List[DataBlock]:
        """解析Excel文件"""
        if not self.can_parse(file_path):
            raise ValueError(f"Unsupported file format: {file_path}")

        ext = file_path.lower().split('.')[-1]

        if ext == 'xls':
            return self._parse_xls(file_path)
        else:
            return self._parse_xlsx(file_path)

    def _parse_xlsx(self, file_path: str) -> List[DataBlock]:
        """解析xlsx文件"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel file parsing. "
                "Install it with: pip install openpyxl"
            )

        blocks = []
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            # 文件打开失败，返回空列表
            print(f"[ExcelParser] 无法打开Excel文件: {e}")
            return blocks

        # 获取缓存目录
        cache_path = self._get_cache_path(file_path)

        for sheet_name in wb.sheetnames:
            try:
                sheet = wb[sheet_name]
                sheet_texts = []

                for row in sheet.iter_rows(values_only=True):
                    # 过滤空行
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    row_text = ' | '.join(row_values)
                    if row_text.strip(' |'):
                        sheet_texts.append(row_text.strip(' |'))

                if sheet_texts:
                    combined_text = "\n".join(sheet_texts)
                    # 安全处理sheet_name中的特殊字符
                    safe_sheet_name = "".join(c if c.isalnum() or c in '_-' else '_' for c in sheet_name)
                    text_cache_path = os.path.join(cache_path, f"sheet_{safe_sheet_name}.txt")
                    try:
                        with open(text_cache_path, 'w', encoding='utf-8') as f:
                            f.write(combined_text)
                    except Exception as e:
                        print(f"[ExcelParser] 保存文本缓存失败: {e}")
                        text_cache_path = None

                    block = DataBlock(
                        block_id=self._generate_block_id(file_path, len(blocks)),
                        modality=ModalityType.TEXT,
                        addr=text_cache_path,
                        file_path=file_path,
                        metadata={
                            'source': 'openpyxl',
                            'sheet_name': sheet_name,
                            'row_count': len(sheet_texts),
                            'text_length': len(combined_text)
                        }
                    )
                    blocks.append(block)
            except Exception as e:
                print(f"[ExcelParser] 解析工作表 {sheet_name} 失败: {e}")

        try:
            wb.close()
        except Exception:
            pass

        return blocks

    def _parse_xls(self, file_path: str) -> List[DataBlock]:
        """解析xls文件（需要LibreOffice转换）"""
        temp_dir = tempfile.gettempdir()
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        xlsx_path = os.path.join(temp_dir, f"{base_name}.xlsx")

        try:
            # 尝试使用LibreOffice转换
            result = subprocess.run(
                ['soffice', '--headless', '--convert-to', 'xlsx',
                 '--outdir', temp_dir, file_path],
                capture_output=True,
                text=True,
                timeout=60
            )

            if result.returncode == 0 and os.path.exists(xlsx_path):
                blocks = self._parse_xlsx(xlsx_path)
                try:
                    os.remove(xlsx_path)
                except Exception:
                    pass
                return blocks
            else:
                error_msg = result.stderr if result.stderr else "未知错误"
                print(f"[ExcelParser] LibreOffice转换失败: {error_msg}")
                return []

        except subprocess.TimeoutExpired:
            print("[ExcelParser] LibreOffice转换超时")
            return []
        except FileNotFoundError:
            print("[ExcelParser] 未找到LibreOffice (soffice命令)，请安装LibreOffice或将.xls转换为.xlsx格式")
            return []
        except Exception as e:
            print(f"[ExcelParser] 解析.xls文件失败: {e}")
            return []